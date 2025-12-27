"""
Gerador de relatórios XLSX para resultados de correção.

Gera planilhas Excel com resultados individuais e resumo estatístico.
"""

import io
import logging
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import ProcessedItem, AnswerKey

logger = logging.getLogger(__name__)


class XLSXGenerator:
    """Gerador de relatórios XLSX para correções."""
    
    # Estilos
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        """Inicializa o gerador."""
        pass
    
    def generate(
        self,
        items: List[ProcessedItem],
        answer_key: AnswerKey,
        job_id: str = "",
        job_created_at: datetime = None
    ) -> bytes:
        """
        Gera arquivo XLSX com resultados da correção.
        
        Args:
            items: Lista de itens processados.
            answer_key: Gabarito usado na correção.
            job_id: ID do job (opcional, para metadados).
            job_created_at: Data de criação do job (opcional).
            
        Returns:
            Bytes do arquivo XLSX.
        """
        wb = Workbook()
        
        # Planilha de resultados
        ws_results = wb.active
        ws_results.title = "Resultados"
        self._create_results_sheet(ws_results, items, answer_key)
        
        # Planilha de resumo estatístico
        ws_summary = wb.create_sheet("Resumo")
        self._create_summary_sheet(ws_summary, items, answer_key, job_id, job_created_at)
        
        # Planilha de gabarito
        ws_answer_key = wb.create_sheet("Gabarito")
        self._create_answer_key_sheet(ws_answer_key, answer_key)
        
        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _create_results_sheet(
        self,
        ws,
        items: List[ProcessedItem],
        answer_key: AnswerKey
    ) -> None:
        """
        Cria a planilha de resultados individuais.
        
        Colunas: Índice, Identificador, Respostas Detectadas, Acertos, Total, Percentual, Status
        """
        # Cabeçalhos
        headers = [
            "Índice",
            "Identificador",
            "Respostas Detectadas",
            "Acertos",
            "Total",
            "Percentual",
            "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Dados
        for row, item in enumerate(items, 2):
            # Índice
            ws.cell(row=row, column=1, value=row - 1).border = self.THIN_BORDER
            
            # Identificador
            ws.cell(row=row, column=2, value=item.identifier or "-").border = self.THIN_BORDER
            
            # Respostas detectadas
            ws.cell(row=row, column=3, value=item.detected_answers or "-").border = self.THIN_BORDER
            
            # Acertos
            ws.cell(row=row, column=4, value=item.correct_count if item.success else "-").border = self.THIN_BORDER
            
            # Total
            ws.cell(row=row, column=5, value=item.total_questions).border = self.THIN_BORDER
            
            # Percentual
            if item.success and item.total_questions > 0:
                percentual = (item.correct_count / item.total_questions) * 100
                cell = ws.cell(row=row, column=6, value=f"{percentual:.1f}%")
            else:
                cell = ws.cell(row=row, column=6, value="-")
            cell.border = self.THIN_BORDER
            
            # Status
            if item.success:
                status_cell = ws.cell(row=row, column=7, value="OK")
                status_cell.fill = self.SUCCESS_FILL
            else:
                error_msg = item.error_message or item.error_code or "Erro"
                status_cell = ws.cell(row=row, column=7, value=f"ERRO: {error_msg}")
                status_cell.fill = self.ERROR_FILL
            status_cell.border = self.THIN_BORDER
        
        # Ajusta largura das colunas
        column_widths = [8, 20, 30, 10, 8, 12, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_summary_sheet(
        self,
        ws,
        items: List[ProcessedItem],
        answer_key: AnswerKey,
        job_id: str,
        job_created_at: datetime
    ) -> None:
        """
        Cria a planilha de resumo estatístico.
        
        Inclui: total de provas, média, mediana, desvio padrão, etc.
        """
        # Calcula estatísticas
        successful_items = [i for i in items if i.success]
        error_items = [i for i in items if not i.success]
        
        scores = [i.correct_count for i in successful_items]
        total_questions = items[0].total_questions if items else 0
        
        # Estatísticas básicas
        total_provas = len(items)
        provas_ok = len(successful_items)
        provas_erro = len(error_items)
        
        if scores:
            media = sum(scores) / len(scores)
            minimo = min(scores)
            maximo = max(scores)
            
            # Mediana
            sorted_scores = sorted(scores)
            n = len(sorted_scores)
            if n % 2 == 0:
                mediana = (sorted_scores[n//2 - 1] + sorted_scores[n//2]) / 2
            else:
                mediana = sorted_scores[n//2]
            
            # Desvio padrão
            if len(scores) > 1:
                variance = sum((x - media) ** 2 for x in scores) / (len(scores) - 1)
                desvio_padrao = variance ** 0.5
            else:
                desvio_padrao = 0
        else:
            media = minimo = maximo = mediana = desvio_padrao = 0
        
        # Título
        ws.cell(row=1, column=1, value="Resumo da Correção").font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Metadados
        row = 3
        metadata = [
            ("ID do Job", job_id or "-"),
            ("Data da Correção", job_created_at.strftime("%d/%m/%Y %H:%M") if job_created_at else "-"),
            ("Total de Questões", total_questions),
        ]
        
        for label, value in metadata:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # Estatísticas
        ws.cell(row=row, column=1, value="Estatísticas").font = Font(bold=True, size=12)
        row += 1
        
        stats = [
            ("Total de Provas", total_provas),
            ("Provas Processadas", provas_ok),
            ("Provas com Erro", provas_erro),
            ("", ""),
            ("Média de Acertos", f"{media:.2f}" if scores else "-"),
            ("Mediana", f"{mediana:.1f}" if scores else "-"),
            ("Mínimo", minimo if scores else "-"),
            ("Máximo", maximo if scores else "-"),
            ("Desvio Padrão", f"{desvio_padrao:.2f}" if scores else "-"),
            ("", ""),
            ("Média Percentual", f"{(media/total_questions*100):.1f}%" if scores and total_questions > 0 else "-"),
        ]
        
        for label, value in stats:
            if label:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Distribuição de notas (se houver dados)
        if scores:
            row += 1
            ws.cell(row=row, column=1, value="Distribuição de Acertos").font = Font(bold=True, size=12)
            row += 1
            
            # Agrupa por faixas de 10%
            faixas = {}
            for score in scores:
                percentual = (score / total_questions) * 100 if total_questions > 0 else 0
                faixa = int(percentual // 10) * 10
                faixa_label = f"{faixa}% - {faixa + 10}%"
                faixas[faixa_label] = faixas.get(faixa_label, 0) + 1
            
            for faixa in sorted(faixas.keys()):
                ws.cell(row=row, column=1, value=faixa)
                ws.cell(row=row, column=2, value=faixas[faixa])
                row += 1
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_answer_key_sheet(self, ws, answer_key: AnswerKey) -> None:
        """
        Cria a planilha com o gabarito usado.
        """
        # Cabeçalhos
        headers = ["Questão", "Resposta"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Dados
        for i, answer in enumerate(answer_key.answers_string, 1):
            ws.cell(row=i + 1, column=1, value=i).border = self.THIN_BORDER
            ws.cell(row=i + 1, column=2, value=answer.upper()).border = self.THIN_BORDER
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
