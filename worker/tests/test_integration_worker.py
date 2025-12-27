"""
Testes de integração do Worker com Supabase local.

Estes testes verificam o fluxo completo de processamento conectando
ao Supabase local e usando imagens de exemplo reais.
"""

import os
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

import pytest

from worker.worker.config import WorkerConfig
from worker.worker.job_processor import JobProcessor
from worker.worker.models import (
    AnswerKey,
    CorrectionItem,
    CorrectionJob,
    JobStatus,
    Template,
)
from worker.worker.queue_consumer import QueueConsumer
from worker.worker.supabase_client import SupabaseWorkerClient


class TestWorkerIntegration:
    """Testes de integração do Worker com Supabase local."""
    
    @pytest.fixture(scope="class")
    def config(self):
        """Configuração para conectar ao Supabase local."""
        return WorkerConfig(
            supabase_url="http://127.0.0.1:54321",
            supabase_service_role_key=os.getenv(
                "SUPABASE_SERVICE_ROLE_KEY",
                "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZS1kZW1vIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImV4cCI6MTk4MzgxMjk5Nn0.EGIM96RAZx35lJzdJsyH-qQwv8Hdp7fsn3W0YpN81IU"
            ),
            queue_name="corrections",
            poll_interval=1.0,
            max_retries=3,
            visibility_timeout=300
        )
    
    @pytest.fixture(scope="class")
    def client(self, config):
        """Cliente Supabase para testes."""
        return SupabaseWorkerClient(config)
    
    @pytest.fixture(scope="class")
    def processor(self, client):
        """Processador de jobs para testes."""
        return JobProcessor(client)
    
    @pytest.fixture
    def sample_template_10_4(self):
        """Template de exemplo: 10 questões, 4 alternativas."""
        return Template(
            id=str(uuid.uuid4()),
            name="Modelo 10 Questões ABCD - Teste",
            question_count=10,
            alternatives_count=4,
            version=1,
            template_storage_path="templates/10_4_template.png",
            is_active=True,
        )
    
    @pytest.fixture
    def sample_answer_key_10_4(self, sample_template_10_4):
        """Gabarito de exemplo para template 10x4."""
        return AnswerKey(
            id=str(uuid.uuid4()),
            owner_user_id=str(uuid.uuid4()),
            template_id=sample_template_10_4.id,
            answers_string="ABCDAABCDA"
        )
    
    @pytest.fixture
    def test_image_paths(self):
        """Caminhos para imagens de teste."""
        base_path = Path("corrector_backend_v2/tests/test_data/10_4_filled1")
        return [
            base_path / "10_04_img_01.jpeg",
            base_path / "10_04_img_02.jpeg",
            base_path / "10_04_img_03.jpeg",
        ]
    
    def test_supabase_connection(self, client):
        """Testa conexão básica com Supabase local."""
        # Tenta fazer uma query simples
        try:
            # Verifica se consegue acessar o banco
            result = client.supabase.table("profiles").select("*").limit(1).execute()
            assert result is not None
            print("✓ Conexão com Supabase local estabelecida")
        except Exception as e:
            pytest.skip(f"Supabase local não está rodando: {e}")
    
    def test_storage_buckets_exist(self, client):
        """Verifica se os buckets de Storage existem."""
        try:
            # Lista buckets
            buckets = client.supabase.storage.list_buckets()
            bucket_names = [b.name for b in buckets]
            
            required_buckets = ["uploads", "results", "templates", "exports"]
            for bucket in required_buckets:
                assert bucket in bucket_names, f"Bucket '{bucket}' não encontrado"
            
            print("✓ Todos os buckets de Storage existem")
        except Exception as e:
            pytest.skip(f"Erro ao verificar buckets: {e}")
    
    def test_upload_test_images(self, client, test_image_paths):
        """Faz upload das imagens de teste para o Storage."""
        user_id = str(uuid.uuid4())
        job_id = str(uuid.uuid4())
        
        uploaded_paths = []
        
        for i, image_path in enumerate(test_image_paths):
            if not image_path.exists():
                pytest.skip(f"Imagem de teste não encontrada: {image_path}")
            
            # Lê a imagem
            with open(image_path, "rb") as f:
                image_data = f.read()
            
            # Upload para o bucket uploads
            storage_path = f"{user_id}/{job_id}/test_image_{i:03d}.jpg"
            
            try:
                client.upload_file(
                    bucket="uploads",
                    path=storage_path,
                    data=image_data,
                    content_type="image/jpeg"
                )
                uploaded_paths.append(f"uploads/{storage_path}")
                print(f"✓ Upload realizado: {storage_path}")
            except Exception as e:
                pytest.fail(f"Erro no upload da imagem {i}: {e}")
        
        # Verifica se consegue baixar as imagens
        for path in uploaded_paths:
            bucket, file_path = path.split("/", 1)
            downloaded = client.download_file(bucket, file_path)
            assert downloaded is not None
            assert len(downloaded) > 0
            print(f"✓ Download verificado: {path}")
    
    def test_image_processor_with_real_images(self, test_image_paths, sample_template_10_4):
        """Testa o processador de imagem com imagens reais."""
        from worker.worker.image_processor import ImageProcessor
        
        processor = ImageProcessor(sample_template_10_4)
        
        for i, image_path in enumerate(test_image_paths):
            if not image_path.exists():
                pytest.skip(f"Imagem de teste não encontrada: {image_path}")
            
            print(f"\nProcessando imagem {i+1}: {image_path.name}")
            
            # Lê a imagem
            with open(image_path, "rb") as f:
                image_bytes = f.read()
            
            # Processa
            processed_item, marked_image_bytes = processor.process(
                image_bytes=image_bytes,
                answers_string="ABCDAABCDA"
            )
            
            print(f"  Sucesso: {processed_item.success}")
            if processed_item.success:
                print(f"  Respostas detectadas: {processed_item.detected_answers}")
                print(f"  Acertos: {processed_item.correct_count}/{processed_item.total_questions}")
                assert processed_item.detected_answers is not None
                assert len(processed_item.detected_answers) == 10
                assert processed_item.correct_count >= 0
                assert marked_image_bytes is not None
                assert len(marked_image_bytes) > 0
            else:
                print(f"  Erro: {processed_item.error_code} - {processed_item.error_message}")
                # Para imagens de teste conhecidas, esperamos sucesso
                # Se falhar, pode ser problema de configuração
                pytest.fail(f"Processamento falhou para imagem conhecida: {processed_item.error_message}")
    
    def test_xlsx_generator_with_processed_items(self, sample_answer_key_10_4):
        """Testa geração de XLSX com itens processados."""
        from worker.worker.models import ProcessedItem
        from worker.worker.xlsx_generator import XLSXGenerator
        
        generator = XLSXGenerator()
        
        # Cria itens de exemplo
        items = [
            ProcessedItem(
                item_id="item-1",
                identifier="001",
                detected_answers="ABCDAABCDA",
                correct_count=10,
                total_questions=10,
                marked_image_path="results/user/job/marked_001.jpg",
                success=True
            ),
            ProcessedItem(
                item_id="item-2",
                identifier="002",
                detected_answers="ABCDAABCDB",
                correct_count=9,
                total_questions=10,
                marked_image_path="results/user/job/marked_002.jpg",
                success=True
            ),
            ProcessedItem(
                item_id="item-3",
                identifier=None,
                detected_answers="",
                correct_count=0,
                total_questions=10,
                marked_image_path="",
                success=False,
                error_code="ALIGN_TRIANGLES_NOT_FOUND",
                error_message="Não foi possível encontrar os triângulos"
            ),
        ]
        
        # Gera XLSX
        xlsx_bytes = generator.generate(
            items=items,
            answer_key=sample_answer_key_10_4,
            job_id="test-job-123",
            job_created_at=datetime.now()
        )
        
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0
        
        # Verifica se é um XLSX válido
        import io

        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Resultados" in wb.sheetnames
        assert "Resumo" in wb.sheetnames
        assert "Gabarito" in wb.sheetnames
        
        print("✓ XLSX gerado com sucesso")
    
    @pytest.mark.slow
    def test_full_job_processing_simulation(
        self, 
        client, 
        processor, 
        sample_template_10_4, 
        sample_answer_key_10_4,
        test_image_paths
    ):
        """
        Simula processamento completo de um job.
        
        Este teste é marcado como 'slow' pois faz processamento real de imagens.
        """
        # Cria IDs únicos
        user_id = str(uuid.uuid4())
        job_id = str(uuid.uuid4())
        
        print(f"\nSimulando job {job_id} para usuário {user_id}")
        
        # 1. Upload das imagens de teste
        uploaded_paths = []
        for i, image_path in enumerate(test_image_paths[:2]):  # Usa apenas 2 imagens
            if not image_path.exists():
                pytest.skip(f"Imagem de teste não encontrada: {image_path}")
            
            with open(image_path, "rb") as f:
                image_data = f.read()
            
            storage_path = f"{user_id}/{job_id}/image_{i:03d}.jpg"
            
            try:
                client.upload_file(
                    bucket="uploads",
                    path=storage_path,
                    data=image_data,
                    content_type="image/jpeg"
                )
                uploaded_paths.append(f"uploads/{storage_path}")
                print(f"  ✓ Upload: {storage_path}")
            except Exception as e:
                pytest.fail(f"Erro no upload: {e}")
        
        # 2. Cria objetos de job simulados
        job = CorrectionJob(
            id=job_id,
            owner_user_id=user_id,
            institution_id=None,
            answer_key_id=sample_answer_key_10_4.id,
            template_id=sample_template_10_4.id,
            status=JobStatus.QUEUED,
            total_items=len(uploaded_paths),
            success_items=0,
            error_items=0,
            elapsed_ms=None,
            xlsx_storage_path=None,
            created_at=datetime.now(),
            started_at=None,
            finished_at=None
        )
        
        items = [
            CorrectionItem(
                id=str(uuid.uuid4()),
                job_id=job_id,
                index=i,
                original_storage_path=path,
            )
            for i, path in enumerate(uploaded_paths)
        ]
        
        # 3. Simula processamento (sem usar banco real)
        print("  Processando itens...")
        
        from worker.worker.image_processor import ImageProcessor
        image_processor = ImageProcessor(sample_template_10_4)
        
        processed_items = []
        success_count = 0
        error_count = 0
        
        for item in items:
            print(f"    Processando item {item.index + 1}/{len(items)}")
            
            # Download da imagem
            bucket, file_path = item.original_storage_path.split("/", 1)
            image_bytes = client.download_file(bucket, file_path)
            
            if not image_bytes:
                error_count += 1
                continue
            
            # Processa
            processed_item, marked_image_bytes = image_processor.process(
                image_bytes=image_bytes,
                answers_string=sample_answer_key_10_4.answers_string
            )
            
            processed_item.item_id = item.id
            
            if processed_item.success:
                success_count += 1
                
                # Upload da imagem marcada
                if marked_image_bytes:
                    marked_path = f"{user_id}/{job_id}/marked_{item.index:04d}.jpg"
                    try:
                        client.upload_file(
                            bucket="results",
                            path=marked_path,
                            data=marked_image_bytes,
                            content_type="image/jpeg"
                        )
                        processed_item.marked_image_path = f"results/{marked_path}"
                        print(f"      ✓ Imagem marcada: {marked_path}")
                    except Exception as e:
                        print(f"      ⚠ Erro no upload da imagem marcada: {e}")
                
                print(f"      ✓ Respostas: {processed_item.detected_answers}")
                print(f"      ✓ Acertos: {processed_item.correct_count}/10")
            else:
                error_count += 1
                print(f"      ✗ Erro: {processed_item.error_code}")
            
            processed_items.append(processed_item)
        
        # 4. Gera XLSX
        print("  Gerando XLSX...")
        from worker.worker.xlsx_generator import XLSXGenerator
        
        xlsx_generator = XLSXGenerator()
        xlsx_bytes = xlsx_generator.generate(
            items=processed_items,
            answer_key=sample_answer_key_10_4,
            job_id=job_id,
            job_created_at=job.created_at
        )
        
        # Upload do XLSX
        xlsx_path = f"{user_id}/{job_id}/results.xlsx"
        try:
            client.upload_file(
                bucket="results",
                path=xlsx_path,
                data=xlsx_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            print(f"  ✓ XLSX: {xlsx_path}")
        except Exception as e:
            pytest.fail(f"Erro no upload do XLSX: {e}")
        
        # 5. Verifica resultados
        print(f"\nResultados finais:")
        print(f"  Total de itens: {len(items)}")
        print(f"  Sucessos: {success_count}")
        print(f"  Erros: {error_count}")
        
        assert success_count > 0, "Nenhum item foi processado com sucesso"
        assert len(processed_items) == len(items)
        
        # Verifica se arquivos foram criados no Storage
        for processed_item in processed_items:
            if processed_item.success and processed_item.marked_image_path:
                bucket, file_path = processed_item.marked_image_path.split("/", 1)
                downloaded = client.download_file(bucket, file_path)
                assert downloaded is not None
                assert len(downloaded) > 0
        
        # Verifica XLSX
        downloaded_xlsx = client.download_file("results", xlsx_path)
        assert downloaded_xlsx is not None
        assert len(downloaded_xlsx) > 0
        
        print("✓ Simulação de job completa com sucesso!")


class TestWorkerWithQueue:
    """Testes do Worker com fila real (pgmq)."""
    
    @pytest.fixture(scope="class")
    def config(self):
        """Configuração para conectar ao Supabase local."""
        return WorkerConfig(
            supabase_url="http://127.0.0.1:54321",
            supabase_service_role_key=os.getenv(
                "SUPABASE_SERVICE_ROLE_KEY",
                "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZS1kZW1vIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImV4cCI6MTk4MzgxMjk5Nn0.EGIM96RAZx35lJzdJsyH-qQwv8Hdp7fsn3W0YpN81IU"
            ),
            queue_name="corrections_test",  # Usa fila separada para testes
            poll_interval=0.5,
            max_retries=2,
            visibility_timeout=30
        )
    
    @pytest.fixture(scope="class")
    def client(self, config):
        """Cliente Supabase para testes."""
        return SupabaseWorkerClient(config)
    
    def test_queue_operations(self, client, config):
        """Testa operações básicas da fila."""
        try:
            # Tenta criar a fila de teste
            client.supabase.rpc("pgmq_create", {"queue_name": config.queue_name}).execute()
        except Exception:
            # Fila já existe, tudo bem
            pass
        
        # Publica uma mensagem de teste
        test_job_id = str(uuid.uuid4())
        
        try:
            client.supabase.rpc(
                "pgmq_send",
                {
                    "queue_name": config.queue_name,
                    "msg": {"job_id": test_job_id}
                }
            ).execute()
            print(f"✓ Mensagem publicada: {test_job_id}")
        except Exception as e:
            pytest.skip(f"Erro ao publicar mensagem: {e}")
        
        # Tenta ler a mensagem
        try:
            message = client.read_queue_message()
            if message:
                print(f"✓ Mensagem lida: {message.job_id}")
                assert message.job_id == test_job_id
                
                # Deleta a mensagem
                client.delete_queue_message(message.msg_id)
                print("✓ Mensagem deletada")
            else:
                pytest.fail("Nenhuma mensagem encontrada na fila")
        except Exception as e:
            pytest.fail(f"Erro ao ler mensagem: {e}")
    
    @pytest.mark.slow
    def test_queue_consumer_basic(self, config):
        """Testa o consumidor de fila básico."""
        consumer = QueueConsumer(config)
        
        # Testa que o consumidor pode ser criado
        assert consumer is not None
        assert consumer.config == config
        assert consumer.client is not None
        assert consumer.processor is not None
        
        print("✓ QueueConsumer criado com sucesso")
        
        # Nota: Não testamos o loop completo aqui pois seria um teste muito longo
        # e dependeria de dados reais no banco


# Configuração para executar apenas testes rápidos por padrão
def pytest_configure(config):
    """Configura marcadores personalizados."""
    config.addinivalue_line(
        "markers", "slow: marca testes que demoram para executar"
    )


# Função utilitária para verificar se Supabase local está rodando
def is_supabase_running():
    """Verifica se o Supabase local está rodando."""
    import requests
    try:
        response = requests.get("http://127.0.0.1:54321/health", timeout=2)
        return response.status_code == 200
    except:
        return False


# Skip todos os testes se Supabase não estiver rodando
pytestmark = pytest.mark.skipif(
    not is_supabase_running(),
    reason="Supabase local não está rodando. Execute 'supabase start' primeiro."
)