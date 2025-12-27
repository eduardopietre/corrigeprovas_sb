"""
Testes para o gerador de XLSX.

Feature: corrige-provas
Validates: Requirements 6.7
"""

import io
import pytest
from datetime import datetime

from openpyxl import load_workbook

from worker.xlsx_generator import XLSXGenerator
from worker.models import ProcessedItem, AnswerKey


class TestXLSXGenerator:
    """Testes unitários para o gerador de XLSX."""
    
    @pytest.fixture
    def generator(self):
        return XLSXGenerator()
    
    @pytest.fixture
    def sample_answer_key(self):
        return AnswerKey(
            id="test-key-1",
            owner_user_id="user-1",
            template_id="template-1",
            answers_string="ABCDAABCDA"
        )
    
    @pytest.fixture
    def sample_items(self):
        return [
            ProcessedItem(
                item_id="item-1",
                identifier="001",
                detected_answers="ABCDAABCDA",
                correct_count=10,
                total_questions=10,
                marked_image_path="results/user-1/job-1/marked_0000.jpg",
                success=True
            ),
            ProcessedItem(
                item_id="item-2",
                identifier="002",
                detected_answers="ABCDAABCDB",
                correct_count=9,
                total_questions=10,
                marked_image_path="results/user-1/job-1/marked_0001.jpg",
                success=True
            ),
            ProcessedItem(
                item_id="item-3",
                identifier=None,
                detected_answers="",
                correct_count=0,
                total_questions=10,
                marked_image_path="",
                success=False,
                error_code="ALIGN_TRIANGLES_NOT_FOUND",
                error_message="Não foi possível encontrar os triângulos de alinhamento"
            ),
        ]
    
    def test_generate_returns_bytes(self, generator, sample_items, sample_answer_key):
        """Deve retornar bytes válidos."""
        result = generator.generate(sample_items, sample_answer_key)
        assert isinstance(result, bytes)
        assert len(result) > 0
    
    def test_generate_valid_xlsx(self, generator, sample_items, sample_answer_key):
        """Deve gerar um XLSX válido que pode ser aberto."""
        xlsx_bytes = generator.generate(sample_items, sample_answer_key)
        
        # Tenta abrir o arquivo
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        
        # Verifica planilhas
        assert "Resultados" in wb.sheetnames
        assert "Resumo" in wb.sheetnames
        assert "Gabarito" in wb.sheetnames
    
    def test_results_sheet_has_correct_headers(self, generator, sample_items, sample_answer_key):
        """A planilha de resultados deve ter os cabeçalhos corretos."""
        xlsx_bytes = generator.generate(sample_items, sample_answer_key)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Resultados"]
        
        expected_headers = [
            "Índice",
            "Identificador",
            "Respostas Detectadas",
            "Acertos",
            "Total",
            "Percentual",
            "Status"
        ]
        
        for col, expected in enumerate(expected_headers, 1):
            assert ws.cell(row=1, column=col).value == expected
    
    def test_results_sheet_has_correct_data(self, generator, sample_items, sample_answer_key):
        """A planilha de resultados deve ter os dados corretos."""
        xlsx_bytes = generator.generate(sample_items, sample_answer_key)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Resultados"]
        
        # Verifica primeira linha de dados (item 1)
        assert ws.cell(row=2, column=2).value == "001"  # Identificador
        assert ws.cell(row=2, column=3).value == "ABCDAABCDA"  # Respostas
        assert ws.cell(row=2, column=4).value == 10  # Acertos
        assert ws.cell(row=2, column=5).value == 10  # Total
        assert ws.cell(row=2, column=7).value == "OK"  # Status
    
    def test_answer_key_sheet_has_all_answers(self, generator, sample_items, sample_answer_key):
        """A planilha de gabarito deve ter todas as respostas."""
        xlsx_bytes = generator.generate(sample_items, sample_answer_key)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Gabarito"]
        
        # Verifica cada resposta
        for i, answer in enumerate(sample_answer_key.answers_string, 1):
            assert ws.cell(row=i + 1, column=1).value == i  # Número da questão
            assert ws.cell(row=i + 1, column=2).value == answer.upper()  # Resposta
    
    def test_summary_sheet_has_statistics(self, generator, sample_items, sample_answer_key):
        """A planilha de resumo deve ter estatísticas."""
        xlsx_bytes = generator.generate(
            sample_items, 
            sample_answer_key,
            job_id="test-job-123",
            job_created_at=datetime(2024, 12, 27, 10, 30)
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Resumo"]
        
        # Verifica que há conteúdo
        assert ws.cell(row=1, column=1).value == "Resumo da Correção"
    
    def test_empty_items_list(self, generator, sample_answer_key):
        """Deve lidar com lista vazia de itens."""
        xlsx_bytes = generator.generate([], sample_answer_key)
        
        # Deve gerar arquivo válido mesmo sem itens
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Resultados" in wb.sheetnames
